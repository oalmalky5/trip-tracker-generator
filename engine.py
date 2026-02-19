# engine.py
# Core generation + Excel export logic for Trip Tracker Generator (Streamlit)

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import random
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

## here we are locking down the expected headers in the generated meetins sheet in a py list.
## needs to be consistent with the template to not allow it to drift and create confusiong for the user
TEMPLATE_HEADERS = [
    "Customer Account Name",
    "Meeting Date",
    "Meeting Time",
    "Meeting City",
    "Meeting Address",
    "East40 Meeting Owner",
    "Primary Contact Name",
    "Primary Contact Email",
    "Meeting Status",
    "Company Description",
]

## here we are lokcing donw the possible status options for meetings in the generated tracker. This is used both in generation and in the template as a dropdown:
STATUS_OPTIONS = ["Proposed", "Tentative", "Confirmed", "Rescheduled", "Cancelled", "Done"]


def suggest_fix(issue: str, field: str = "", context: str = "") -> str:
    """
    Heuristic suggestions used in the Data Issues tab.
    issue: human-readable message
    field: the field name (if available)
    context: optional extra context (not required)
    """
    t = (issue or "").lower()
    f = (field or "").lower()
    c = (context or "").lower()

    blob = " ".join([t, f, c])

    if "email" in blob:
        return "Add a valid email address for the primary contact in CRM (Contacts export)."
    if "address" in blob or "location" in blob or "hq" in blob:
        return "Add the meeting address / HQ address in CRM (Accounts export)."
    if "owner" in blob:
        return "Assign an account owner in CRM so meetings can be distributed."
    if "description" in blob:
        return "Add a short company description in CRM (optional)."
    if "duplicate" in blob:
        return "De-duplicate the record in CRM or confirm which record should be used."
    if "missing" in blob or "blank" in blob:
        return "Fill the missing value in CRM and re-export."
    if "no matching contacts" in blob:
        return "Ensure contacts are associated with the correct company (website/company name match) and re-export."
    return "Review the CRM export and correct the value."


@dataclass(frozen=True)
class TripConfig:
    trip_name: str
    start_date: date
    end_date: date
    meetings: int = 12
    city: str = "Riyadh"
    owners: Tuple[str, ...] = ("Jason", "Meshari")
    seed: int = 42


@dataclass
class Issue:
    severity: str   # "BLOCKER" or "WARNING"
    entity: str     # "account" / "contact" / "meeting"
    entity_id: str  # e.g. Company ID or Person ID
    field: str
    message: str


def _safe_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def load_accounts_contacts(accounts_path: Path, contacts_path: Optional[Path]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Loads the case-study exports (Excel).
    contacts_path can be None (app supports accounts-only flow).
    """
    accounts = pd.read_excel(accounts_path)

    if contacts_path is None:
        contacts = pd.DataFrame()
    else:
        contacts = pd.read_excel(contacts_path)

    return accounts, contacts


def pick_accounts(accounts: pd.DataFrame, n: int, seed: int, city: Optional[str] = None) -> pd.DataFrame:
    """
    Picks N accounts.
    If city is provided and 'HQ City' exists, we prefer accounts with that city (when possible).
    """
    rnd = random.Random(seed)
    df = accounts.copy()

    if city and "HQ City" in df.columns:
        df_city = df[df["HQ City"].fillna("").astype(str).str.lower().str.contains(city.lower())]
        if len(df_city) >= n:
            df = df_city

    if len(df) <= n:
        return df.sample(n=min(n, len(df)), random_state=seed).reset_index(drop=True)

    idx = rnd.sample(list(df.index), k=n)
    return df.loc[idx].reset_index(drop=True)


def select_primary_contact_for_account(
    account_row: pd.Series,
    contacts: pd.DataFrame,
    issues: List[Issue],
) -> Tuple[str, str]:
    """
    Return (name, email).

    Strategy:
    1) Use account's 'Primary Contact Email' if present
    2) Else, if contacts provided:
        a) Match by Primary Company Website == account Website
        b) Else match by Primary Company == account Companies
       Choose first contact with an email. If account Primary Contact name exists, prefer exact match.

    If no contacts provided or no match found, return ("","") and emit WARNING.
    """
    company_id = _safe_str(account_row.get("Company ID", ""))
    acct_primary_name = _safe_str(account_row.get("Primary Contact", ""))
    acct_primary_email = _safe_str(account_row.get("Primary Contact Email", ""))
    acct_company = _safe_str(account_row.get("Companies", ""))
    acct_website = _safe_str(account_row.get("Website", ""))

    if acct_primary_email:
        return (acct_primary_name or "", acct_primary_email)

    if contacts is None or contacts.empty:
        issues.append(Issue(
            severity="WARNING",
            entity="account",
            entity_id=company_id or acct_company or "(unknown)",
            field="Primary Contact Email",
            message="Contacts export not provided; primary contact left blank."
        ))
        return ("", "")

    # Guard columns existence
    has_site = "Primary Company Website" in contacts.columns
    has_company = "Primary Company" in contacts.columns
    has_email = "Email" in contacts.columns
    has_people = "People" in contacts.columns

    if not has_email or not has_people or (not has_site and not has_company):
        issues.append(Issue(
            severity="WARNING",
            entity="account",
            entity_id=company_id or acct_company or "(unknown)",
            field="Primary Contact Email",
            message="Contacts export missing expected columns; primary contact left blank."
        ))
        return ("", "")

    cands = contacts.copy()

    if acct_website and has_site:
        cands = cands[cands["Primary Company Website"].fillna("").astype(str).str.lower() == acct_website.lower()]

    if cands.empty and acct_company and has_company:
        cands = contacts[contacts["Primary Company"].fillna("").astype(str).str.lower() == acct_company.lower()]

    if cands.empty:
        issues.append(Issue(
            severity="WARNING",
            entity="account",
            entity_id=company_id or acct_company or "(unknown)",
            field="Primary Contact Email",
            message="No matching contacts found; primary contact left blank."
        ))
        return ("", "")

    # Prefer contacts with email
    cands = cands.copy()
    cands["_email"] = cands["Email"].apply(_safe_str)
    cands = cands[cands["_email"].astype(bool)]
    if cands.empty:
        issues.append(Issue(
            severity="WARNING",
            entity="account",
            entity_id=company_id or acct_company or "(unknown)",
            field="Primary Contact Email",
            message="Matching contacts exist but none have an email; primary contact left blank."
        ))
        return ("", "")

    # Prefer exact name match if account has a primary contact name
    if acct_primary_name:
        norm_target = _norm(acct_primary_name)
        cands["_name"] = cands["People"].apply(_safe_str)
        exact = cands[cands["_name"].apply(_norm) == norm_target]
        if not exact.empty:
            row = exact.iloc[0]
            return (_safe_str(row.get("People", "")), _safe_str(row.get("Email", "")))

    row = cands.iloc[0]
    return (_safe_str(row.get("People", "")), _safe_str(row.get("Email", "")))


def generate_schedule(cfg: TripConfig, n: int) -> List[Tuple[date, str]]:
    """
    Returns list of (meeting_date, meeting_time).
    Distributes meetings across days; times are within business hours (09:00-17:30), 30-min increments.
    """
    rnd = random.Random(cfg.seed)

    days = (cfg.end_date - cfg.start_date).days + 1
    if days <= 0:
        raise ValueError("end_date must be on or after start_date")

    per_day = [0] * days
    for i in range(n):
        per_day[i % days] += 1

    slots: List[Tuple[date, str]] = []
    for d, count in enumerate(per_day):
        day = cfg.start_date + timedelta(days=d)
        possible = [f"{h:02d}:{m:02d}" for h in range(9, 18) for m in (0, 30)]  # 09:00 -> 17:30
        rnd.shuffle(possible)
        chosen = sorted(possible[:count])
        for t in chosen:
            slots.append((day, t))

    return slots


def build_meetings_df(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    cfg: TripConfig,
) -> Tuple[pd.DataFrame, List[Issue], Dict]:
    issues: List[Issue] = []

    picked = pick_accounts(accounts, cfg.meetings, cfg.seed, city=cfg.city)
    schedule = generate_schedule(cfg, len(picked))

    rows: List[Dict] = []
    owners = list(cfg.owners) if cfg.owners else ["Owner"]
    rnd = random.Random(cfg.seed)

    for i, acct in picked.iterrows():
        meeting_date, meeting_time = schedule[i]
        owner = owners[i % len(owners)]
        status = rnd.choice(STATUS_OPTIONS)

        acct_id = _safe_str(acct.get("Company ID", "(unknown)"))
        acct_name = _safe_str(acct.get("Companies", ""))
        desc = _safe_str(acct.get("Description", ""))

        addr1 = _safe_str(acct.get("HQ Address Line 1", ""))
        addr2 = _safe_str(acct.get("HQ Address Line 2", ""))
        hq_city = _safe_str(acct.get("HQ City", ""))
        city_val = hq_city or cfg.city
        address = ", ".join([p for p in [addr1, addr2, city_val] if p])

        contact_name, contact_email = select_primary_contact_for_account(acct, contacts, issues)

        if not acct_name:
            issues.append(Issue(
                severity="BLOCKER",
                entity="account",
                entity_id=acct_id,
                field="Companies",
                message="Missing account name."
            ))

        if not address:
            issues.append(Issue(
                severity="WARNING",
                entity="account",
                entity_id=acct_id,
                field="HQ Address",
                message="Missing HQ address; meeting address left blank."
            ))

        if not contact_email:
            issues.append(Issue(
                severity="WARNING",
                entity="contact",
                entity_id=acct_id,
                field="Email",
                message="Primary contact email missing."
            ))

        rows.append({
            "Customer Account Name": acct_name,
            "Meeting Date": meeting_date.strftime("%b %d, %Y"),
            "Meeting Time": meeting_time,
            "Meeting City": cfg.city,
            "Meeting Address": address,
            "East40 Meeting Owner": owner,
            "Primary Contact Name": contact_name,
            "Primary Contact Email": contact_email,
            "Meeting Status": status,
            "Company Description": desc,
        })

    meetings_df = pd.DataFrame(rows, columns=TEMPLATE_HEADERS)

    industries = picked["Primary Industry Group"] if "Primary Industry Group" in picked.columns else None
    ind_counts = industries.fillna("(blank)").astype(str).value_counts().to_dict() if industries is not None else {}

    stats = {
        "accounts_loaded": int(len(accounts)),
        "contacts_loaded": int(len(contacts)) if contacts is not None else 0,
        "meetings": int(len(meetings_df)),
        "days": int((cfg.end_date - cfg.start_date).days + 1),
        "owner_counts": meetings_df["East40 Meeting Owner"].value_counts().to_dict() if not meetings_df.empty else {},
        "status_counts": meetings_df["Meeting Status"].value_counts().to_dict() if not meetings_df.empty else {},
        "industry_counts": ind_counts,
    }

    return meetings_df, issues, stats


def _auto_fit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def export_excel(
    template_path: Path,
    output_path: Path,
    cfg: TripConfig,
    meetings_df: pd.DataFrame,
    contacts_df: pd.DataFrame,
    issues: List[Issue],
    stats: Dict,
    run_log_lines: List[str],
) -> Path:
    """
    Writes the generated tracker based on the provided template.
    Produces:
      - Trip Overview
      - Meetings
      - Contacts Directory
      - Summary
      - Data Issues (with Suggested Fix)
    """
    wb = load_workbook(template_path)

    # Rename main sheet to Meetings (assume first sheet is the template sheet)
    main = wb[wb.sheetnames[0]]
    main.title = "Meetings"

    # Clear existing rows except header
    if main.max_row > 1:
        main.delete_rows(2, main.max_row - 1)

    # Write meetings rows
    for r_idx, row in enumerate(meetings_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            main.cell(row=r_idx, column=c_idx, value=value)

    main.freeze_panes = "A2"
    main.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_HEADERS))}{main.max_row}"

    # Trip Overview
    if "Trip Overview" in wb.sheetnames:
        del wb["Trip Overview"]
    ov = wb.create_sheet("Trip Overview", 0)
    ov["A1"] = "Trip"
    ov["B1"] = cfg.trip_name
    ov["A2"] = "Dates"
    ov["B2"] = f"{cfg.start_date.isoformat()} to {cfg.end_date.isoformat()}"
    ov["A3"] = "City"
    ov["B3"] = cfg.city
    ov["A4"] = "Meetings"
    ov["B4"] = stats.get("meetings", 0)

    ov["A6"] = "Run Log"
    ov["A6"].font = Font(bold=True)
    for i, line in enumerate(run_log_lines, start=7):
        ov[f"A{i}"] = line

    # Summary
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    sm = wb.create_sheet("Summary", 1)
    sm["A1"] = "Summary"
    sm["A1"].font = Font(bold=True)

    def _write_counts(start_row: int, title: str, counts: Dict[str, int]) -> int:
        sm[f"A{start_row}"] = title
        sm[f"A{start_row}"].font = Font(bold=True)
        r = start_row + 1
        sm[f"A{r}"] = "Category"
        sm[f"B{r}"] = "Count"
        sm[f"A{r}"].font = Font(bold=True)
        sm[f"B{r}"].font = Font(bold=True)
        r += 1
        for k, v in (counts or {}).items():
            sm[f"A{r}"] = k
            sm[f"B{r}"] = int(v)
            r += 1
        return r + 1

    r = 3
    r = _write_counts(r, "Meetings by Status", stats.get("status_counts", {}))
    r = _write_counts(r, "Meetings by Owner", stats.get("owner_counts", {}))
    r = _write_counts(r, "Accounts by Industry Group", stats.get("industry_counts", {}))

    # Contacts Directory (for selected accounts)
    if "Contacts Directory" in wb.sheetnames:
        del wb["Contacts Directory"]
    cd = wb.create_sheet("Contacts Directory")

    if contacts_df is None or contacts_df.empty:
        cd["A1"] = "No contacts provided or no matching contacts found."
    else:
        contact_cols = [
            c for c in [
                "People", "Email", "Phone", "Primary Position", "Primary Company",
                "Primary Company Type", "City", "Country/Territory/Region", "LinkedIn URL"
            ]
            if c in contacts_df.columns
        ]
        for c_idx, col in enumerate(contact_cols, start=1):
            cd.cell(row=1, column=c_idx, value=col).font = Font(bold=True)
        for r_idx, row in enumerate(contacts_df[contact_cols].fillna("").itertuples(index=False), start=2):
            for c_idx, v in enumerate(row, start=1):
                cd.cell(row=r_idx, column=c_idx, value=v)

        cd.freeze_panes = "A2"
        if contact_cols:
            cd.auto_filter.ref = f"A1:{get_column_letter(len(contact_cols))}{cd.max_row}"

    # Data Issues (with Suggested Fix)
    if "Data Issues" in wb.sheetnames:
        del wb["Data Issues"]
    di = wb.create_sheet("Data Issues")
    di_headers = ["Severity", "Entity", "Entity ID", "Field", "Message", "Suggested Fix"]
    for c_idx, h in enumerate(di_headers, start=1):
        di.cell(row=1, column=c_idx, value=h).font = Font(bold=True)

    for r_idx, iss in enumerate(issues, start=2):
        di.cell(row=r_idx, column=1, value=iss.severity)
        di.cell(row=r_idx, column=2, value=iss.entity)
        di.cell(row=r_idx, column=3, value=iss.entity_id)
        di.cell(row=r_idx, column=4, value=iss.field)
        di.cell(row=r_idx, column=5, value=iss.message)
        di.cell(row=r_idx, column=6, value=suggest_fix(iss.message, iss.field, iss.entity))

    di.freeze_panes = "A2"
    di.auto_filter.ref = f"A1:{get_column_letter(len(di_headers))}{di.max_row}"

    # Fit columns
    for ws in [ov, sm, cd, di, main]:
        _auto_fit_columns(ws)

    wb.save(output_path)
    return output_path